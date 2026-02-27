"""Генератор Excel-отчётов на основе pandas."""

from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from database.models import Booking, Equipment
from utils.logger import logger


async def generate_report(
    session: AsyncSession,
    days: Optional[int],
    category_id: Optional[int] = None,
    user_id: Optional[int] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    bot=None,
) -> Optional[Path]:
    """
    Сгенерировать Excel-отчёт по бронированиям с опциональными фильтрами.

    days=None означает произвольный диапазон (используются start_date/end_date).
    Возвращает путь к файлу или None при ошибке.
    """
    try:
        now = datetime.now(timezone.utc)

        if days is not None:
            date_from = now - timedelta(days=days)
            date_to = now
        elif start_date and end_date:
            date_from = start_date
            date_to = end_date
        else:
            date_from = now - timedelta(days=30)
            date_to = now

        query = (
            select(Booking)
            .where(Booking.created_at >= date_from)
            .where(Booking.created_at <= date_to)
            .options(
                selectinload(Booking.user),
                selectinload(Booking.equipment)
            )
            .order_by(Booking.created_at.desc())
        )

        if category_id is not None:
            query = query.join(Equipment, Booking.equipment_id == Equipment.id).where(
                Equipment.category_id == category_id
            )
        if user_id is not None:
            query = query.where(Booking.user_id == user_id)

        result = await session.execute(query)
        bookings = list(result.scalars().all())

        if not bookings:
            logger.info("No bookings found for report")
            return None

        data = []
        for booking in bookings:
            if booking.start_time and booking.end_time:
                duration = booking.end_time - booking.start_time
                duration_hours = duration.total_seconds() / 3600
            else:
                duration_hours = 0

            created_str = booking.created_at.strftime("%Y-%m-%d %H:%M") if booking.created_at else ""
            start_str = booking.start_time.strftime("%Y-%m-%d %H:%M") if booking.start_time else ""
            end_str = booking.end_time.strftime("%Y-%m-%d %H:%M") if booking.end_time else ""
            confirmed_str = booking.confirmed_at.strftime("%Y-%m-%d %H:%M") if booking.confirmed_at else ""
            completed_str = booking.completed_at.strftime("%Y-%m-%d %H:%M") if booking.completed_at else ""

            status_map = {
                "pending": "Ожидает",
                "active": "Активна",
                "completed": "Завершена",
                "cancelled": "Отменена",
                "expired": "Истекла",
                "maintenance": "Тех. обслуживание",
            }
            status_ru = status_map.get(booking.status, booking.status)

            data.append({
                "ID брони": booking.id,
                "Статус": status_ru,
                "Сотрудник": booking.user.full_name,
                "Telegram ID": booking.user.telegram_id,
                "Телефон": booking.user.phone_number or "",
                "Оборудование": booking.equipment.name,
                "Категория": booking.equipment.category,
                "Дата создания": created_str,
                "Начало брони": start_str,
                "Конец брони": end_str,
                "Длительность (ч)": round(duration_hours, 1),
                "Подтверждена": confirmed_str,
                "Завершена": completed_str,
                "Просрочка": "Да" if booking.is_overdue else "Нет",
                "Фото начало": len(booking.photos_start) if booking.photos_start else 0,
                "Фото конец": len(booking.photos_end) if booking.photos_end else 0,
            })

        df = pd.DataFrame(data)

        reports_dir = Path("reports/files")
        reports_dir.mkdir(parents=True, exist_ok=True)

        timestamp = now.strftime("%Y%m%d_%H%M%S")
        parts = ["booking_report"]
        if days is not None:
            parts.append(f"{days}days")
        else:
            parts.append(f"{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}")
        if category_id:
            parts.append(f"cat{category_id}")
        if user_id:
            parts.append(f"user{user_id}")
        parts.append(timestamp)
        file_path = reports_dir / f"{'_'.join(parts)}.xlsx"

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Бронирования')

            workbook = writer.book
            worksheet = writer.sheets['Бронирования']

            # Авто-ширина колонок (макс. 50 символов)
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                max_length = min(max_length, 50)
                worksheet.column_dimensions[chr(65 + idx)].width = max_length

            filter_lines = []
            if days is not None:
                filter_lines.append(f"Период: последние {days} дней")
            else:
                filter_lines.append(
                    f"Период: {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
                )
            if category_id is not None:
                cat_names = set(b.equipment.category for b in bookings if b.equipment)
                filter_lines.append(f"Категория: {', '.join(cat_names) or f'ID {category_id}'}")
            if user_id is not None:
                user_names = set(b.user.full_name for b in bookings if b.user)
                filter_lines.append(f"Сотрудник: {', '.join(user_names) or f'ID {user_id}'}")

            summary_data = {
                "Метрика": [
                    "Фильтры",
                    "Всего броней",
                    "Активных",
                    "Завершённых",
                    "Отменённых",
                    "Истекших",
                    "Просроченных",
                    "Уникальных сотрудников",
                    "Уникальных объектов",
                ],
                "Значение": [
                    "; ".join(filter_lines),
                    len(bookings),
                    sum(1 for b in bookings if b.status == "active"),
                    sum(1 for b in bookings if b.status == "completed"),
                    sum(1 for b in bookings if b.status == "cancelled"),
                    sum(1 for b in bookings if b.status == "expired"),
                    sum(1 for b in bookings if b.is_overdue),
                    len(set(b.user_id for b in bookings)),
                    len(set(b.equipment_id for b in bookings)),
                ]
            }

            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, index=False, sheet_name='Сводка')

            summary_sheet = writer.sheets['Сводка']
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 15

            bookings_with_photos = [
                b for b in bookings
                if (b.photos_start and len(b.photos_start) > 0)
                or (b.photos_end and len(b.photos_end) > 0)
            ]
            if bookings_with_photos:
                workbook.create_sheet('Фото броней')
                photo_sheet = workbook['Фото броней']
                photo_sheet.column_dimensions['A'].width = 10
                photo_sheet.column_dimensions['B'].width = 25
                photo_sheet.column_dimensions['C'].width = 25
                photo_sheet.column_dimensions['D'].width = 8
                photo_sheet.column_dimensions['E'].width = 30

                photo_sheet.append(["ID брони", "Сотрудник", "Оборудование", "Тип", "Файл"])
                current_row = 2

                for booking in bookings_with_photos:
                    all_photos = []
                    for path in (booking.photos_start or []):
                        all_photos.append(("Начало", path))
                    for path in (booking.photos_end or []):
                        all_photos.append(("Конец", path))

                    for photo_type, photo_path in all_photos:
                        photo_sheet.cell(current_row, 1, booking.id)
                        photo_sheet.cell(current_row, 2, booking.user.full_name)
                        photo_sheet.cell(current_row, 3, booking.equipment.name)
                        photo_sheet.cell(current_row, 4, photo_type)
                        photo_sheet.cell(current_row, 5, photo_path)

                        # Встраиваем изображение, если это локальный файл
                        local_file = Path(photo_path)
                        if local_file.exists():
                            try:
                                with PILImage.open(local_file) as img:
                                    orig_w, orig_h = img.size

                                max_w, max_h = 200, 150
                                ratio = min(max_w / orig_w, max_h / orig_h, 1.0)
                                display_w = int(orig_w * ratio)
                                display_h = int(orig_h * ratio)

                                xl_img = XLImage(str(local_file))
                                xl_img.width = display_w
                                xl_img.height = display_h
                                photo_sheet.row_dimensions[current_row].height = display_h * 0.75 + 5
                                photo_sheet.add_image(xl_img, f"F{current_row}")
                            except Exception as img_err:
                                logger.warning(f"Could not embed image {photo_path}: {img_err}")

                        current_row += 1

        logger.info(
            f"Generated report: {file_path.name}, "
            f"{len(bookings)} bookings"
        )

        return file_path

    except Exception as e:
        logger.error(f"Error generating report: {e}", exc_info=True)
        return None
